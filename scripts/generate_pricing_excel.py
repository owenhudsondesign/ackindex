#!/usr/bin/env python3
"""
Generate a dynamic Excel pricing model for government contracting.
All costs cascade from input assumptions - change inputs and everything updates.
"""

import openpyxl
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.chart import BarChart, Reference

def create_pricing_workbook():
    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    section_font = Font(bold=True, size=11)
    section_fill = PatternFill(start_color="D6DCE5", end_color="D6DCE5", fill_type="solid")
    input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Yellow for editable
    output_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Green for calculated
    currency_format = '"$"#,##0.00'
    percent_format = '0.0%'
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ==========================================
    # SHEET 1: INPUTS (All editable values)
    # ==========================================
    ws_inputs = wb.active
    ws_inputs.title = "Inputs"

    # Column widths
    ws_inputs.column_dimensions['A'].width = 35
    ws_inputs.column_dimensions['B'].width = 18
    ws_inputs.column_dimensions['C'].width = 15
    ws_inputs.column_dimensions['D'].width = 40

    row = 1

    # Title
    ws_inputs['A1'] = "ACKINDEX GOVERNMENT CONTRACT PRICING MODEL"
    ws_inputs['A1'].font = Font(bold=True, size=16)
    ws_inputs.merge_cells('A1:D1')
    row = 3

    # Instructions
    ws_inputs['A3'] = "📝 INSTRUCTIONS: Edit YELLOW cells to update all calculations"
    ws_inputs['A3'].font = Font(italic=True, size=10)
    ws_inputs.merge_cells('A3:D3')
    row = 5

    # --- VOLUME ASSUMPTIONS ---
    ws_inputs[f'A{row}'] = "VOLUME ASSUMPTIONS"
    ws_inputs[f'A{row}'].font = section_font
    ws_inputs[f'A{row}'].fill = section_fill
    ws_inputs.merge_cells(f'A{row}:D{row}')
    row += 1

    inputs_start = row
    volume_inputs = [
        ("Monthly Footage (minutes)", 5000, "min/month", "Benchmark: 5,000 min = 83.3 hours"),
        ("Monthly Queries", 5000, "queries/month", "User questions per month"),
        ("Storage per Minute (MB)", 125, "MB", "Compressed video storage"),
        ("Monthly Bandwidth (GB)", 500, "GB", "CDN streaming to users"),
    ]

    for label, value, unit, note in volume_inputs:
        ws_inputs[f'A{row}'] = label
        ws_inputs[f'B{row}'] = value
        ws_inputs[f'B{row}'].fill = input_fill
        ws_inputs[f'B{row}'].number_format = '#,##0'
        ws_inputs[f'C{row}'] = unit
        ws_inputs[f'D{row}'] = note
        row += 1

    row += 1

    # --- INFRASTRUCTURE COSTS ---
    ws_inputs[f'A{row}'] = "INFRASTRUCTURE UNIT COSTS"
    ws_inputs[f'A{row}'].font = section_font
    ws_inputs[f'A{row}'].fill = section_fill
    ws_inputs.merge_cells(f'A{row}:D{row}')
    row += 1

    infra_inputs = [
        ("Supabase Pro (monthly)", 25.00, "$/month", "PostgreSQL + pgvector database"),
        ("Vercel Pro (monthly)", 20.00, "$/month", "Next.js hosting"),
        ("Upstash Redis (monthly)", 0.00, "$/month", "Free tier caching"),
        ("Sentry (monthly)", 0.00, "$/month", "Free tier monitoring"),
    ]

    for label, value, unit, note in infra_inputs:
        ws_inputs[f'A{row}'] = label
        ws_inputs[f'B{row}'] = value
        ws_inputs[f'B{row}'].fill = input_fill
        ws_inputs[f'B{row}'].number_format = currency_format
        ws_inputs[f'C{row}'] = unit
        ws_inputs[f'D{row}'] = note
        row += 1

    row += 1

    # --- STORAGE COSTS ---
    ws_inputs[f'A{row}'] = "STORAGE UNIT COSTS"
    ws_inputs[f'A{row}'].font = section_font
    ws_inputs[f'A{row}'].fill = section_fill
    ws_inputs.merge_cells(f'A{row}:D{row}')
    row += 1

    storage_inputs = [
        ("Bunny.net Storage", 0.005, "$/GB/month", "Video file storage"),
        ("Bunny.net Bandwidth", 0.01, "$/GB", "CDN streaming"),
    ]

    for label, value, unit, note in storage_inputs:
        ws_inputs[f'A{row}'] = label
        ws_inputs[f'B{row}'] = value
        ws_inputs[f'B{row}'].fill = input_fill
        ws_inputs[f'B{row}'].number_format = '"$"#,##0.0000'
        ws_inputs[f'C{row}'] = unit
        ws_inputs[f'D{row}'] = note
        row += 1

    row += 1

    # --- TRANSCRIPTION COSTS ---
    ws_inputs[f'A{row}'] = "TRANSCRIPTION UNIT COSTS"
    ws_inputs[f'A{row}'].font = section_font
    ws_inputs[f'A{row}'].fill = section_fill
    ws_inputs.merge_cells(f'A{row}:D{row}')
    row += 1

    transcription_inputs = [
        ("AssemblyAI (per minute)", 0.0222, "$/min", "Primary - best quality"),
        ("Gladia Nova-2 (per minute)", 0.0043, "$/min", "Alternative - budget"),
    ]

    for label, value, unit, note in transcription_inputs:
        ws_inputs[f'A{row}'] = label
        ws_inputs[f'B{row}'] = value
        ws_inputs[f'B{row}'].fill = input_fill
        ws_inputs[f'B{row}'].number_format = '"$"#,##0.0000'
        ws_inputs[f'C{row}'] = unit
        ws_inputs[f'D{row}'] = note
        row += 1

    row += 1

    # --- AI COSTS ---
    ws_inputs[f'A{row}'] = "AI SERVICE UNIT COSTS"
    ws_inputs[f'A{row}'].font = section_font
    ws_inputs[f'A{row}'].fill = section_fill
    ws_inputs.merge_cells(f'A{row}:D{row}')
    row += 1

    ai_inputs = [
        ("Voyage AI Embeddings", 0.06, "$/1M tokens", "Semantic search vectors"),
        ("Claude Haiku Input", 0.80, "$/1M tokens", "Simple queries"),
        ("Claude Haiku Output", 4.00, "$/1M tokens", "Simple responses"),
        ("Claude Sonnet Input", 3.00, "$/1M tokens", "Complex queries"),
        ("Claude Sonnet Output", 15.00, "$/1M tokens", "Complex responses"),
        ("Haiku Query Percentage", 0.80, "%", "% of queries using Haiku"),
        ("Tokens per Query (input)", 800, "tokens", "Average input tokens"),
        ("Tokens per Query (output)", 400, "tokens", "Average output tokens"),
        ("Tokens per Minute (embedding)", 150, "tokens", "For new content"),
    ]

    for label, value, unit, note in ai_inputs:
        ws_inputs[f'A{row}'] = label
        ws_inputs[f'B{row}'] = value
        ws_inputs[f'B{row}'].fill = input_fill
        if "%" in unit:
            ws_inputs[f'B{row}'].number_format = percent_format
        elif "$" in unit:
            ws_inputs[f'B{row}'].number_format = '"$"#,##0.00'
        else:
            ws_inputs[f'B{row}'].number_format = '#,##0'
        ws_inputs[f'C{row}'] = unit
        ws_inputs[f'D{row}'] = note
        row += 1

    row += 1

    # --- SETUP COSTS ---
    ws_inputs[f'A{row}'] = "ONE-TIME SETUP COSTS"
    ws_inputs[f'A{row}'].font = section_font
    ws_inputs[f'A{row}'].fill = section_fill
    ws_inputs.merge_cells(f'A{row}:D{row}')
    row += 1

    setup_inputs = [
        ("Professional Services Rate", 150.00, "$/hour", "Integration & customization"),
        ("Professional Services Hours", 40, "hours", "Estimated setup time"),
        ("Training Rate", 250.00, "$/hour", "Staff training"),
        ("Training Hours", 8, "hours", "Admin + user training"),
    ]

    for label, value, unit, note in setup_inputs:
        ws_inputs[f'A{row}'] = label
        ws_inputs[f'B{row}'] = value
        ws_inputs[f'B{row}'].fill = input_fill
        if "$" in unit:
            ws_inputs[f'B{row}'].number_format = currency_format
        else:
            ws_inputs[f'B{row}'].number_format = '#,##0'
        ws_inputs[f'C{row}'] = unit
        ws_inputs[f'D{row}'] = note
        row += 1

    row += 1

    # --- MARKUP ---
    ws_inputs[f'A{row}'] = "MARKUP & MARGIN SETTINGS"
    ws_inputs[f'A{row}'].font = section_font
    ws_inputs[f'A{row}'].fill = section_fill
    ws_inputs.merge_cells(f'A{row}:D{row}')
    row += 1

    markup_inputs = [
        ("GSA Schedule Markup", 0.15, "%", "Standard 15% government markup"),
        ("Target Gross Margin", 0.50, "%", "Target profit margin on pricing"),
    ]

    for label, value, unit, note in markup_inputs:
        ws_inputs[f'A{row}'] = label
        ws_inputs[f'B{row}'] = value
        ws_inputs[f'B{row}'].fill = input_fill
        ws_inputs[f'B{row}'].number_format = percent_format
        ws_inputs[f'C{row}'] = unit
        ws_inputs[f'D{row}'] = note
        row += 1

    # ==========================================
    # SHEET 2: COST BREAKDOWN (All formulas)
    # ==========================================
    ws_costs = wb.create_sheet("Cost Breakdown")

    ws_costs.column_dimensions['A'].width = 35
    ws_costs.column_dimensions['B'].width = 18
    ws_costs.column_dimensions['C'].width = 18
    ws_costs.column_dimensions['D'].width = 18
    ws_costs.column_dimensions['E'].width = 35

    row = 1
    ws_costs['A1'] = "MONTHLY COST BREAKDOWN"
    ws_costs['A1'].font = Font(bold=True, size=14)
    ws_costs.merge_cells('A1:E1')
    row = 3

    # Headers
    headers = ["Component", "Monthly Cost", "Annual Cost", "% of Total", "Formula Reference"]
    for col, header in enumerate(headers, 1):
        cell = ws_costs.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    row += 1

    # Infrastructure Section
    ws_costs[f'A{row}'] = "INFRASTRUCTURE"
    ws_costs[f'A{row}'].font = section_font
    ws_costs[f'A{row}'].fill = section_fill
    ws_costs.merge_cells(f'A{row}:E{row}')
    row += 1

    infra_rows = [
        ("Supabase Pro", "=Inputs!B10", "Database"),
        ("Vercel Pro", "=Inputs!B11", "Hosting"),
        ("Upstash Redis", "=Inputs!B12", "Caching"),
        ("Sentry", "=Inputs!B13", "Monitoring"),
    ]

    infra_start = row
    for label, formula, note in infra_rows:
        ws_costs[f'A{row}'] = label
        ws_costs[f'B{row}'] = formula
        ws_costs[f'B{row}'].number_format = currency_format
        ws_costs[f'B{row}'].fill = output_fill
        ws_costs[f'C{row}'] = f"=B{row}*12"
        ws_costs[f'C{row}'].number_format = currency_format
        ws_costs[f'C{row}'].fill = output_fill
        ws_costs[f'E{row}'] = note
        row += 1
    infra_end = row - 1

    ws_costs[f'A{row}'] = "Infrastructure Subtotal"
    ws_costs[f'A{row}'].font = Font(bold=True)
    ws_costs[f'B{row}'] = f"=SUM(B{infra_start}:B{infra_end})"
    ws_costs[f'B{row}'].number_format = currency_format
    ws_costs[f'B{row}'].font = Font(bold=True)
    ws_costs[f'B{row}'].fill = output_fill
    ws_costs[f'C{row}'] = f"=B{row}*12"
    ws_costs[f'C{row}'].number_format = currency_format
    ws_costs[f'C{row}'].font = Font(bold=True)
    ws_costs[f'C{row}'].fill = output_fill
    infra_subtotal_row = row
    row += 2

    # Storage Section
    ws_costs[f'A{row}'] = "STORAGE"
    ws_costs[f'A{row}'].font = section_font
    ws_costs[f'A{row}'].fill = section_fill
    ws_costs.merge_cells(f'A{row}:E{row}')
    row += 1

    storage_start = row
    # Video Storage: (minutes * MB per min / 1024) * price per GB
    ws_costs[f'A{row}'] = "Video Storage"
    ws_costs[f'B{row}'] = "=(Inputs!B6*Inputs!B8/1024)*Inputs!B16"
    ws_costs[f'B{row}'].number_format = currency_format
    ws_costs[f'B{row}'].fill = output_fill
    ws_costs[f'C{row}'] = f"=B{row}*12"
    ws_costs[f'C{row}'].number_format = currency_format
    ws_costs[f'C{row}'].fill = output_fill
    ws_costs[f'E{row}'] = "=(minutes × MB/min ÷ 1024) × $/GB"
    row += 1

    # Bandwidth
    ws_costs[f'A{row}'] = "Video Bandwidth (CDN)"
    ws_costs[f'B{row}'] = "=Inputs!B9*Inputs!B17"
    ws_costs[f'B{row}'].number_format = currency_format
    ws_costs[f'B{row}'].fill = output_fill
    ws_costs[f'C{row}'] = f"=B{row}*12"
    ws_costs[f'C{row}'].number_format = currency_format
    ws_costs[f'C{row}'].fill = output_fill
    ws_costs[f'E{row}'] = "=GB × $/GB"
    row += 1
    storage_end = row - 1

    ws_costs[f'A{row}'] = "Storage Subtotal"
    ws_costs[f'A{row}'].font = Font(bold=True)
    ws_costs[f'B{row}'] = f"=SUM(B{storage_start}:B{storage_end})"
    ws_costs[f'B{row}'].number_format = currency_format
    ws_costs[f'B{row}'].font = Font(bold=True)
    ws_costs[f'B{row}'].fill = output_fill
    ws_costs[f'C{row}'] = f"=B{row}*12"
    ws_costs[f'C{row}'].number_format = currency_format
    ws_costs[f'C{row}'].font = Font(bold=True)
    ws_costs[f'C{row}'].fill = output_fill
    storage_subtotal_row = row
    row += 2

    # Transcription Section
    ws_costs[f'A{row}'] = "TRANSCRIPTION"
    ws_costs[f'A{row}'].font = section_font
    ws_costs[f'A{row}'].fill = section_fill
    ws_costs.merge_cells(f'A{row}:E{row}')
    row += 1

    ws_costs[f'A{row}'] = "AssemblyAI Transcription"
    ws_costs[f'B{row}'] = "=Inputs!B6*Inputs!B20"
    ws_costs[f'B{row}'].number_format = currency_format
    ws_costs[f'B{row}'].fill = output_fill
    ws_costs[f'C{row}'] = f"=B{row}*12"
    ws_costs[f'C{row}'].number_format = currency_format
    ws_costs[f'C{row}'].fill = output_fill
    ws_costs[f'E{row}'] = "=minutes × $/min"
    transcription_row = row
    row += 2

    # AI Section
    ws_costs[f'A{row}'] = "AI SERVICES"
    ws_costs[f'A{row}'].font = section_font
    ws_costs[f'A{row}'].fill = section_fill
    ws_costs.merge_cells(f'A{row}:E{row}')
    row += 1

    ai_start = row

    # Embeddings: minutes * tokens/min / 1M * $/1M tokens
    ws_costs[f'A{row}'] = "Embeddings (New Content)"
    ws_costs[f'B{row}'] = "=Inputs!B6*Inputs!B32/1000000*Inputs!B24"
    ws_costs[f'B{row}'].number_format = currency_format
    ws_costs[f'B{row}'].fill = output_fill
    ws_costs[f'C{row}'] = f"=B{row}*12"
    ws_costs[f'C{row}'].number_format = currency_format
    ws_costs[f'C{row}'].fill = output_fill
    ws_costs[f'E{row}'] = "=min × tokens/min ÷ 1M × $/1M"
    row += 1

    # Haiku queries: queries * haiku% * tokens / 1M * $/1M (input + output)
    ws_costs[f'A{row}'] = "Claude Haiku (Simple Queries)"
    ws_costs[f'B{row}'] = "=Inputs!B7*Inputs!B29*((Inputs!B30/1000000*Inputs!B25)+(Inputs!B31/1000000*Inputs!B26))"
    ws_costs[f'B{row}'].number_format = currency_format
    ws_costs[f'B{row}'].fill = output_fill
    ws_costs[f'C{row}'] = f"=B{row}*12"
    ws_costs[f'C{row}'].number_format = currency_format
    ws_costs[f'C{row}'].fill = output_fill
    ws_costs[f'E{row}'] = "=queries × haiku% × token costs"
    row += 1

    # Sonnet queries
    ws_costs[f'A{row}'] = "Claude Sonnet (Complex Queries)"
    ws_costs[f'B{row}'] = "=Inputs!B7*(1-Inputs!B29)*((Inputs!B30/1000000*Inputs!B27)+(Inputs!B31/1000000*Inputs!B28))"
    ws_costs[f'B{row}'].number_format = currency_format
    ws_costs[f'B{row}'].fill = output_fill
    ws_costs[f'C{row}'] = f"=B{row}*12"
    ws_costs[f'C{row}'].number_format = currency_format
    ws_costs[f'C{row}'].fill = output_fill
    ws_costs[f'E{row}'] = "=queries × sonnet% × token costs"
    row += 1
    ai_end = row - 1

    ws_costs[f'A{row}'] = "AI Services Subtotal"
    ws_costs[f'A{row}'].font = Font(bold=True)
    ws_costs[f'B{row}'] = f"=SUM(B{ai_start}:B{ai_end})"
    ws_costs[f'B{row}'].number_format = currency_format
    ws_costs[f'B{row}'].font = Font(bold=True)
    ws_costs[f'B{row}'].fill = output_fill
    ws_costs[f'C{row}'] = f"=B{row}*12"
    ws_costs[f'C{row}'].number_format = currency_format
    ws_costs[f'C{row}'].font = Font(bold=True)
    ws_costs[f'C{row}'].fill = output_fill
    ai_subtotal_row = row
    row += 2

    # TOTAL
    ws_costs[f'A{row}'] = "TOTAL MONTHLY OPERATIONS"
    ws_costs[f'A{row}'].font = Font(bold=True, size=12)
    ws_costs[f'B{row}'] = f"=B{infra_subtotal_row}+B{storage_subtotal_row}+B{transcription_row}+B{ai_subtotal_row}"
    ws_costs[f'B{row}'].number_format = currency_format
    ws_costs[f'B{row}'].font = Font(bold=True, size=12)
    ws_costs[f'B{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ws_costs[f'C{row}'] = f"=B{row}*12"
    ws_costs[f'C{row}'].number_format = currency_format
    ws_costs[f'C{row}'].font = Font(bold=True, size=12)
    ws_costs[f'C{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    total_row = row
    row += 2

    # Per-minute cost
    ws_costs[f'A{row}'] = "Cost Per Minute of Footage"
    ws_costs[f'A{row}'].font = Font(bold=True)
    ws_costs[f'B{row}'] = f"=B{total_row}/Inputs!B6"
    ws_costs[f'B{row}'].number_format = '"$"#,##0.0000'
    ws_costs[f'B{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ws_costs[f'E{row}'] = "=Total ÷ minutes"
    cost_per_min_row = row
    row += 2

    # One-time costs
    ws_costs[f'A{row}'] = "ONE-TIME SETUP COSTS"
    ws_costs[f'A{row}'].font = section_font
    ws_costs[f'A{row}'].fill = section_fill
    ws_costs.merge_cells(f'A{row}:E{row}')
    row += 1

    ws_costs[f'A{row}'] = "Professional Services"
    ws_costs[f'B{row}'] = "=Inputs!B35*Inputs!B36"
    ws_costs[f'B{row}'].number_format = currency_format
    ws_costs[f'B{row}'].fill = output_fill
    ws_costs[f'E{row}'] = "=rate × hours"
    row += 1

    ws_costs[f'A{row}'] = "Training"
    ws_costs[f'B{row}'] = "=Inputs!B37*Inputs!B38"
    ws_costs[f'B{row}'].number_format = currency_format
    ws_costs[f'B{row}'].fill = output_fill
    ws_costs[f'E{row}'] = "=rate × hours"
    row += 1

    ws_costs[f'A{row}'] = "Setup Total"
    ws_costs[f'A{row}'].font = Font(bold=True)
    ws_costs[f'B{row}'] = f"=B{row-2}+B{row-1}"
    ws_costs[f'B{row}'].number_format = currency_format
    ws_costs[f'B{row}'].font = Font(bold=True)
    ws_costs[f'B{row}'].fill = output_fill
    setup_total_row = row
    row += 2

    # With GSA markup
    ws_costs[f'A{row}'] = "WITH GSA MARKUP"
    ws_costs[f'A{row}'].font = section_font
    ws_costs[f'A{row}'].fill = section_fill
    ws_costs.merge_cells(f'A{row}:E{row}')
    row += 1

    ws_costs[f'A{row}'] = "Monthly Operations (with markup)"
    ws_costs[f'B{row}'] = f"=B{total_row}*(1+Inputs!B41)"
    ws_costs[f'B{row}'].number_format = currency_format
    ws_costs[f'B{row}'].fill = output_fill
    row += 1

    ws_costs[f'A{row}'] = "Setup (with markup)"
    ws_costs[f'B{row}'] = f"=B{setup_total_row}*(1+Inputs!B41)"
    ws_costs[f'B{row}'].number_format = currency_format
    ws_costs[f'B{row}'].fill = output_fill

    # ==========================================
    # SHEET 3: PRICING TIERS
    # ==========================================
    ws_pricing = wb.create_sheet("Pricing Tiers")

    ws_pricing.column_dimensions['A'].width = 30
    ws_pricing.column_dimensions['B'].width = 15
    ws_pricing.column_dimensions['C'].width = 15
    ws_pricing.column_dimensions['D'].width = 15
    ws_pricing.column_dimensions['E'].width = 15
    ws_pricing.column_dimensions['F'].width = 15
    ws_pricing.column_dimensions['G'].width = 15
    ws_pricing.column_dimensions['H'].width = 15

    row = 1
    ws_pricing['A1'] = "CONTRACT PRICING TIERS"
    ws_pricing['A1'].font = Font(bold=True, size=14)
    ws_pricing.merge_cells('A1:H1')
    row = 3

    # Input your prices here
    ws_pricing[f'A{row}'] = "SET YOUR PRICES (edit yellow cells)"
    ws_pricing[f'A{row}'].font = section_font
    ws_pricing[f'A{row}'].fill = section_fill
    ws_pricing.merge_cells(f'A{row}:H{row}')
    row += 1

    # Headers
    price_headers = ["Tier", "Minutes/Month", "Your Price", "Cost", "Gross Profit", "Margin %", "Overage $/min", "Setup Fee"]
    for col, header in enumerate(price_headers, 1):
        cell = ws_pricing.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    row += 1

    tiers = [
        ("Basic", 2500, 299),
        ("Standard", 5000, 399),
        ("Professional", 10000, 699),
        ("Enterprise", 25000, 1499),
    ]

    tier_start = row
    for tier_name, minutes, suggested_price in tiers:
        ws_pricing[f'A{row}'] = tier_name
        ws_pricing[f'B{row}'] = minutes
        ws_pricing[f'B{row}'].number_format = '#,##0'

        # Your price (editable)
        ws_pricing[f'C{row}'] = suggested_price
        ws_pricing[f'C{row}'].fill = input_fill
        ws_pricing[f'C{row}'].number_format = currency_format

        # Cost (calculated from cost breakdown)
        ws_pricing[f'D{row}'] = f"='Cost Breakdown'!B{cost_per_min_row}*B{row}+'Cost Breakdown'!B{infra_subtotal_row}"
        ws_pricing[f'D{row}'].number_format = currency_format
        ws_pricing[f'D{row}'].fill = output_fill

        # Gross Profit
        ws_pricing[f'E{row}'] = f"=C{row}-D{row}"
        ws_pricing[f'E{row}'].number_format = currency_format
        ws_pricing[f'E{row}'].fill = output_fill

        # Margin %
        ws_pricing[f'F{row}'] = f"=IF(C{row}>0,E{row}/C{row},0)"
        ws_pricing[f'F{row}'].number_format = percent_format
        ws_pricing[f'F{row}'].fill = output_fill

        # Overage rate (editable)
        ws_pricing[f'G{row}'] = 0.15 - (0.02 * (tiers.index((tier_name, minutes, suggested_price))))
        ws_pricing[f'G{row}'].fill = input_fill
        ws_pricing[f'G{row}'].number_format = '"$"#,##0.00'

        # Setup fee
        ws_pricing[f'H{row}'] = f"='Cost Breakdown'!B{setup_total_row}*(1+Inputs!B41)"
        ws_pricing[f'H{row}'].number_format = currency_format
        ws_pricing[f'H{row}'].fill = output_fill

        row += 1
    tier_end = row - 1

    row += 2

    # Annual calculations
    ws_pricing[f'A{row}'] = "ANNUAL CONTRACT VALUES"
    ws_pricing[f'A{row}'].font = section_font
    ws_pricing[f'A{row}'].fill = section_fill
    ws_pricing.merge_cells(f'A{row}:H{row}')
    row += 1

    annual_headers = ["Tier", "Monthly Fee", "Annual Fee", "Setup", "Year 1 Total", "Year 2+ Annual", "Year 1 Margin", "Break-even Months"]
    for col, header in enumerate(annual_headers, 1):
        cell = ws_pricing.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    row += 1

    for i, (tier_name, minutes, _) in enumerate(tiers):
        tier_row = tier_start + i
        ws_pricing[f'A{row}'] = tier_name
        ws_pricing[f'B{row}'] = f"=C{tier_row}"
        ws_pricing[f'B{row}'].number_format = currency_format
        ws_pricing[f'C{row}'] = f"=C{tier_row}*12"
        ws_pricing[f'C{row}'].number_format = currency_format
        ws_pricing[f'D{row}'] = f"=H{tier_row}"
        ws_pricing[f'D{row}'].number_format = currency_format
        ws_pricing[f'E{row}'] = f"=C{row}+D{row}"
        ws_pricing[f'E{row}'].number_format = currency_format
        ws_pricing[f'E{row}'].fill = output_fill
        ws_pricing[f'F{row}'] = f"=C{row}"
        ws_pricing[f'F{row}'].number_format = currency_format
        ws_pricing[f'G{row}'] = f"=(C{row}-(D{tier_row}*12))/E{row}"
        ws_pricing[f'G{row}'].number_format = percent_format
        ws_pricing[f'G{row}'].fill = output_fill
        ws_pricing[f'H{row}'] = f"=IF(E{tier_row}>0,D{row}/E{tier_row},0)"
        ws_pricing[f'H{row}'].number_format = '0.0'
        ws_pricing[f'H{row}'].fill = output_fill
        row += 1

    # ==========================================
    # SHEET 4: GRANT BUDGET
    # ==========================================
    ws_grant = wb.create_sheet("Grant Budget")

    ws_grant.column_dimensions['A'].width = 35
    ws_grant.column_dimensions['B'].width = 15
    ws_grant.column_dimensions['C'].width = 15
    ws_grant.column_dimensions['D'].width = 15
    ws_grant.column_dimensions['E'].width = 15
    ws_grant.column_dimensions['F'].width = 30

    row = 1
    ws_grant['A1'] = "3-YEAR GRANT BUDGET TEMPLATE"
    ws_grant['A1'].font = Font(bold=True, size=14)
    ws_grant.merge_cells('A1:F1')
    row = 3

    # Select tier
    ws_grant[f'A{row}'] = "Select Minutes Tier:"
    ws_grant[f'B{row}'] = 5000
    ws_grant[f'B{row}'].fill = input_fill
    ws_grant[f'B{row}'].number_format = '#,##0'
    ws_grant[f'C{row}'] = "min/month"
    selected_tier_row = row
    row += 2

    # Headers
    grant_headers = ["Line Item", "Year 1", "Year 2", "Year 3", "3-Year Total", "Notes"]
    for col, header in enumerate(grant_headers, 1):
        cell = ws_grant.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    row += 1

    # Setup
    ws_grant[f'A{row}'] = "Platform Setup & Integration"
    ws_grant[f'B{row}'] = f"='Cost Breakdown'!B{setup_total_row}*(1+Inputs!B41)"
    ws_grant[f'B{row}'].number_format = currency_format
    ws_grant[f'B{row}'].fill = output_fill
    ws_grant[f'C{row}'] = 0
    ws_grant[f'C{row}'].number_format = currency_format
    ws_grant[f'D{row}'] = 0
    ws_grant[f'D{row}'].number_format = currency_format
    ws_grant[f'E{row}'] = f"=SUM(B{row}:D{row})"
    ws_grant[f'E{row}'].number_format = currency_format
    ws_grant[f'E{row}'].fill = output_fill
    ws_grant[f'F{row}'] = "One-time"
    row += 1

    # Annual operations - using cost per minute * selected tier minutes + infrastructure
    ws_grant[f'A{row}'] = "Annual Platform Operations"
    ws_grant[f'B{row}'] = f"=(('Cost Breakdown'!B{cost_per_min_row}*B{selected_tier_row})+'Cost Breakdown'!B{infra_subtotal_row})*12*(1+Inputs!B41)"
    ws_grant[f'B{row}'].number_format = currency_format
    ws_grant[f'B{row}'].fill = output_fill
    ws_grant[f'C{row}'] = f"=B{row}"
    ws_grant[f'C{row}'].number_format = currency_format
    ws_grant[f'C{row}'].fill = output_fill
    ws_grant[f'D{row}'] = f"=B{row}"
    ws_grant[f'D{row}'].number_format = currency_format
    ws_grant[f'D{row}'].fill = output_fill
    ws_grant[f'E{row}'] = f"=SUM(B{row}:D{row})"
    ws_grant[f'E{row}'].number_format = currency_format
    ws_grant[f'E{row}'].fill = output_fill
    ws_grant[f'F{row}'] = f"={selected_tier_row} min/month tier"
    ops_row = row
    row += 1

    # Contingency
    ws_grant[f'A{row}'] = "Contingency (10%)"
    ws_grant[f'B{row}'] = f"=(B{row-2}+B{row-1})*0.1"
    ws_grant[f'B{row}'].number_format = currency_format
    ws_grant[f'B{row}'].fill = output_fill
    ws_grant[f'C{row}'] = f"=C{row-1}*0.1"
    ws_grant[f'C{row}'].number_format = currency_format
    ws_grant[f'C{row}'].fill = output_fill
    ws_grant[f'D{row}'] = f"=D{row-1}*0.1"
    ws_grant[f'D{row}'].number_format = currency_format
    ws_grant[f'D{row}'].fill = output_fill
    ws_grant[f'E{row}'] = f"=SUM(B{row}:D{row})"
    ws_grant[f'E{row}'].number_format = currency_format
    ws_grant[f'E{row}'].fill = output_fill
    row += 1

    # Total
    ws_grant[f'A{row}'] = "GRANT TOTAL"
    ws_grant[f'A{row}'].font = Font(bold=True, size=12)
    for col in ['B', 'C', 'D', 'E']:
        ws_grant[f'{col}{row}'] = f"=SUM({col}{row-3}:{col}{row-1})"
        ws_grant[f'{col}{row}'].number_format = currency_format
        ws_grant[f'{col}{row}'].font = Font(bold=True, size=12)
        ws_grant[f'{col}{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # ==========================================
    # SHEET 5: ROI CALCULATOR
    # ==========================================
    ws_roi = wb.create_sheet("ROI Calculator")

    ws_roi.column_dimensions['A'].width = 35
    ws_roi.column_dimensions['B'].width = 18
    ws_roi.column_dimensions['C'].width = 18
    ws_roi.column_dimensions['D'].width = 18
    ws_roi.column_dimensions['E'].width = 30

    row = 1
    ws_roi['A1'] = "ROI ANALYSIS FOR MUNICIPALITIES"
    ws_roi['A1'].font = Font(bold=True, size=14)
    ws_roi.merge_cells('A1:E1')
    row = 3

    # Inputs
    ws_roi[f'A{row}'] = "MUNICIPALITY ASSUMPTIONS (edit yellow)"
    ws_roi[f'A{row}'].font = section_font
    ws_roi[f'A{row}'].fill = section_fill
    ws_roi.merge_cells(f'A{row}:E{row}')
    row += 1

    roi_inputs = [
        ("Staff Hours Saved per Month", 40, "hours", "Answering constituent inquiries"),
        ("Average Staff Hourly Cost", 50, "$/hour", "Including benefits"),
        ("Selected Pricing Tier", 5000, "min/month", "Match to Pricing Tiers sheet"),
    ]

    for label, value, unit, note in roi_inputs:
        ws_roi[f'A{row}'] = label
        ws_roi[f'B{row}'] = value
        ws_roi[f'B{row}'].fill = input_fill
        if "$" in unit:
            ws_roi[f'B{row}'].number_format = currency_format
        else:
            ws_roi[f'B{row}'].number_format = '#,##0'
        ws_roi[f'C{row}'] = unit
        ws_roi[f'D{row}'] = note
        row += 1

    hours_row = row - 3
    rate_row = row - 2
    tier_row = row - 1
    row += 1

    # Calculations
    ws_roi[f'A{row}'] = "CALCULATED VALUES"
    ws_roi[f'A{row}'].font = section_font
    ws_roi[f'A{row}'].fill = section_fill
    ws_roi.merge_cells(f'A{row}:E{row}')
    row += 1

    ws_roi[f'A{row}'] = "Monthly Labor Savings"
    ws_roi[f'B{row}'] = f"=B{hours_row}*B{rate_row}"
    ws_roi[f'B{row}'].number_format = currency_format
    ws_roi[f'B{row}'].fill = output_fill
    monthly_savings_row = row
    row += 1

    ws_roi[f'A{row}'] = "Annual Labor Savings"
    ws_roi[f'B{row}'] = f"=B{monthly_savings_row}*12"
    ws_roi[f'B{row}'].number_format = currency_format
    ws_roi[f'B{row}'].fill = output_fill
    annual_savings_row = row
    row += 1

    ws_roi[f'A{row}'] = "Annual Platform Cost"
    ws_roi[f'B{row}'] = f"=(('Cost Breakdown'!B{cost_per_min_row}*B{tier_row})+'Cost Breakdown'!B{infra_subtotal_row})*12*(1+Inputs!B41)"
    ws_roi[f'B{row}'].number_format = currency_format
    ws_roi[f'B{row}'].fill = output_fill
    annual_cost_row = row
    row += 1

    ws_roi[f'A{row}'] = "Net Annual Benefit"
    ws_roi[f'B{row}'] = f"=B{annual_savings_row}-B{annual_cost_row}"
    ws_roi[f'B{row}'].number_format = currency_format
    ws_roi[f'B{row}'].font = Font(bold=True)
    ws_roi[f'B{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    row += 1

    ws_roi[f'A{row}'] = "ROI Percentage"
    ws_roi[f'B{row}'] = f"=IF(B{annual_cost_row}>0,B{row-1}/B{annual_cost_row},0)"
    ws_roi[f'B{row}'].number_format = '0%'
    ws_roi[f'B{row}'].font = Font(bold=True, size=14)
    ws_roi[f'B{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    row += 1

    ws_roi[f'A{row}'] = "Payback Period"
    ws_roi[f'B{row}'] = f"=IF(B{monthly_savings_row}>0,B{annual_cost_row}/12/B{monthly_savings_row},0)"
    ws_roi[f'B{row}'].number_format = '0.0" months"'
    ws_roi[f'B{row}'].fill = output_fill

    # Save workbook
    output_path = "/Users/owenhudson/ackindex/AckIndex_Government_Pricing_Model.xlsx"
    wb.save(output_path)
    print(f"✅ Created: {output_path}")
    print("\nSheets created:")
    print("  1. Inputs - All editable values (yellow cells)")
    print("  2. Cost Breakdown - Calculated costs with formulas")
    print("  3. Pricing Tiers - Set your prices, see margins")
    print("  4. Grant Budget - 3-year budget template")
    print("  5. ROI Calculator - Municipality ROI analysis")

if __name__ == "__main__":
    create_pricing_workbook()
